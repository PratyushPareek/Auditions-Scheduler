from openpyxl import load_workbook

#Define
minRow = 2
maxRow = 10

#Functions

def implement(no,so,t,s):
    print("In implement\n")
    if(s==minFree[so][t]):
        minFree[so][t]+= 1
        while(slotSheet.cell(row=minFree[so][t],column=so*2+t+1).value):
            minFree[so][t]+= 1
    print(str(so)+": "+str(t)+": "+str(minFree[so][t])+"\n")
    return s

def isOk(no,so,t,s,slots):
    print("In isOk\n")
    if slotSheet.cell(row=s,column=so*2+t+1).value:
        return False
    for slot in slots:
        if (slot==s or slot-1==s or slot+1==s):
            return False
    return True

def allot(no,so,t,slots):
    print("In allot\n")
    s = minFree[so][t]
    while(not isOk(no,so,t,s,slots)):
        s+= 1
    return implement(no,so,t,s)

#Main Function

#Opening Files
print("Atleast I started\n")
inFile = load_workbook(filename="InSheet.xlsx")
inSheet = inFile.active
print("Loaded insheet\n")
studFile = load_workbook(filename="StudSheet.xlsx")
studSheet = studFile.active
print("Loaded studsheet\n")
slotFile = load_workbook(filename="SlotSheet.xlsx")
slotSheet = slotFile.active
print("Loaded slotsheet\n")
socFile = []
socSheets=[[]]*6
for i in range(0,6):
    socFile.append(load_workbook(filename="SocSheets\\"+str(i)+".xlsx"))
    socSheets[i].append(socFile[i]["Thread A"])
    socSheets[i].append(socFile[i]["Thread B"])
    print("Loaded "+str(i)+"th socsheet\n")

minFree = [[2 for i in range(2)] for j in range(6)]
minStud = 2
minSocs = [[2 for i in range(2)] for j in range(6)]

for freshie in inSheet.iter_rows(min_row=minRow,max_row=maxRow,min_col=1,max_col=9):
    #Fetch data
    print("In\n")
    rollNo = freshie[0].value
    name = freshie[1].value
    contact = freshie[2].value
    status=[]
    for i in range(3,9):
        status.append(bool(freshie[i].value))
    print(status)
    print("\n")
    slots=[]
    print("Got it\n")
    for i in range(0,6):
        if(status[i]):
            print("IN "+str(i)+"\n")
            t=-1
            s=-1
            print(minFree)
            if(minFree[i][0]>minFree[i][1]):
                t=1
            else:
                t=0
            s = allot(rollNo,i,t,slots)
            slots.append(s)
            #Slot Entry
            print("Found\n")
            slotSheet.cell(row=s,column=i*2+t+1).value = rollNo
            #Student Entry
            studSheet.cell(row=minStud, column=1).value = rollNo
            studSheet.cell(row=minStud, column=2+i*2).value = s
            studSheet.cell(row=minStud, column=3+i*2).value = t
            #Society Entry
            socSheets[i][t].cell(row=s, column=1).value = rollNo
            socSheets[i][t].cell(row=s, column=2).value = name
            socSheets[i][t].cell(row=s, column=3).value = contact
            print("Done\n")
    minStud += 1
    inFile.save("InSheet.xlsx")
    studFile.save("StudSheet.xlsx")
    slotFile.save("SlotSheet.xlsx")
    for i in range(0,6):
        socFile[i].save("SocSheets\\"+str(i)+".xlsx")
